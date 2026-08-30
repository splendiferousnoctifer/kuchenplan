"""Import küchenplan_26.xlsx into SQLite, swapping Kaiserschmarrn → Palatschinken."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

from .db import connect, init_db, portion_equivalents

DAY_INDEX = {
    "Sonntag": 0,
    "Montag": 1,
    "Dienstag": 2,
    "Mittwoch": 3,
    "Donnerstag": 4,
    "Freitag": 5,
    "Samstag": 6,
}

# GuteKüche "Einfaches Palatschinkenrezept" for 5 portions, scaled to 40
# https://www.gutekueche.at/einfaches-palatschinkenrezept-rezept-4278
PALATSCHINKEN_LINES = [
    # ingredient, unit, amount_for_40, notes
    ("Mehl", "kg", 2.0, "250 g × 8 (Weizenmehl glatt)"),
    ("Milch", "l", 4.0, "500 ml × 8"),
    ("Eier", "stk", 16.0, "2 Stk × 8"),
    ("Salz", "pkg", 1.0, "1 Prise × 8 → 1 Packung fürs Lager"),
    ("Öl", "el", 8.0, "1 Schuss/Pfanne ≈ 1 EL × 8"),
    ("Staubzucker", "kg", 0.08, "1 EL × 8 ≈ 80 g"),
    ("Marmelade", "kg", 0.64, "4 EL × 8 ≈ 640 g"),
]

PALATSCHINKEN_SOURCE = (
    "https://www.gutekueche.at/einfaches-palatschinkenrezept-rezept-4278"
)

# Veggie-only lines: sheet scaled them by full camp (42). Re-base as per-vegetarian
# amounts. Spinatknödel fixed at 10 total for 2 vegetarians (user).
# Other batch totals kept ≈ original recipe_amount for the veggie group.
VEGGIE_LINE_FIXES = [
    # (recipe, ingredient, amount_per_veggie, recipe_amount_total, note)
    (
        "Knödel",
        "Spinatknödel",
        5.0,
        10.0,
        "veggie-only; 10 stk for 2 vegetarians",
    ),
    (
        "Putenschnitzel",
        "Vegane Schnitzel",
        2.0,
        4.0,
        "veggie-only; was 4 stk in 40-portion recipe",
    ),
    (
        "Fleischlaibchen",
        "Gemüselaibchen",
        2.0,
        4.0,
        "veggie-only; was 4 stk in 40-portion recipe",
    ),
]

# Absolute batch overrides from prior-year notes (for yield_portions=40 recipes)
BATCH_OVERRIDES = [
    # (recipe, ingredient, new_batch_amount, note)
    ("Bosner", "Frankfurter", 80.0, "80 statt 100"),
    ("Bosner", "Weckerl", 80.0, "80 statt 100 (matched to Frankfurter)"),
    (
        "Kartoffelsalat",
        "Zwiebel",
        0.4,
        "was 4 kg for 40 — treated as typo, set to 0.4 kg",
    ),
]

# Unit/amount rewrites for ambiguous pkg items
# (recipe, ingredient, unit, batch, yield, audience, note)
UNIT_REWRITES = [
    ("Spaghetti", "Tomaten Passiert", "g", 3000.0, 40.0, "all",
     "was 3 pkg; 1 pkg = 1 l ≈ 1000 g"),
    ("Spaghetti", "Tomaten Gewürfelt", "g", 1000.0, 40.0, "all",
     "was 2 pkg; counted as 500 g packs"),
    ("Spaghetti", "Tomatenmark", "Tube", 3.0, 40.0, "all",
     "was pkg → Tuben"),
    ("Knödel", "Kraut", "kg", 4.0, 40.0, "all",
     "Sackerl 500 g; was 8 pkg → 4 kg"),
    ("Grillabend", "Haloumi", "g", 1600.0, 2.0, "veggie",
     "was 8 pkg; 200 g/Packung → 1600 g for 2 veggies"),
]

# Former placeholder zeros → sensible batch amounts for 40 portions
# (recipe, ingredient, unit, batch_amount, note)
ZERO_FILLS = [
    ("Putenschnitzel", "Milch", "l", 1.0, "1 l for panieren"),
    ("Grüner Salat", "Olivenöl", "l", 0.15, "dressing ~4 ml/person"),
    ("Grüner Salat", "Essig", "l", 0.1, "dressing"),
    ("Grüner Salat", "Senf", "kg", 0.05, "dressing"),
    ("Grüner Salat", "Pfeffer", "pkg", 1.0, "seasoning for salad"),
    ("Grüner Salat", "Salz", "pkg", 1.0, "seasoning for salad"),
    ("Gurkensalat", "Olivenöl", "l", 0.08, "dressing"),
    ("Gurkensalat", "Essig", "l", 0.15, "vinegar-forward salad"),
    ("Kartoffelsalat", "Olivenöl", "l", 0.25, "dressing"),
    ("Kartoffelsalat", "Essig", "l", 0.25, "dressing"),
    ("Karottensalat", "Olivenöl", "l", 0.05, "dressing"),
    ("Karottensalat", "Tafelessig", "l", 0.08, "dressing"),
    ("Spaghetti", "Olivenöl", "l", 0.05, "for sauce"),
    ("Spaghetti", "Ketchup", "kg", 0.3, "optional for sauce/serving"),
    ("Fleischlaibchen", "Ketchup", "kg", 0.25, "to serve"),
    ("Fleischlaibchen", "Knoblauchgranulat", "pkg", 1.0, "seasoning"),
    ("Fleischlaibchen", "Muskat-Nuss", "pkg", 1.0, "seasoning"),
]


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        if v.startswith("="):
            # simple D/E or 23-7
            m = re.fullmatch(r"=(\d+(?:\.\d+)?)\s*/\s*\$?([A-Z]+\$?)?(\d+(?:\.\d+)?)", v)
            if m:
                return float(m.group(1)) / float(m.group(3))
            m = re.fullmatch(r"=(\d+)\s*-\s*(\d+)", v)
            if m:
                return float(m.group(1)) - float(m.group(2))
            return None
        try:
            return float(v)
        except ValueError:
            return None
    return None


def _cached_dummy(val):
    if not isinstance(val, str) or "DUMMYFUNCTION" not in val:
        return val
    m = re.search(r'COMPUTED_VALUE""\"\),(.*)\)\s*$', val, re.S)
    if not m:
        return val
    raw = m.group(1).strip()
    if raw.startswith('"') and raw.endswith('"'):
        return raw[1:-1]
    try:
        return float(raw)
    except ValueError:
        return raw


def _split_recipes(cell: str) -> list[str]:
    return [p.strip() for p in str(cell).split(",") if p.strip()]


def _get_or_create_ingredient(conn, name: str, unit: str | None, cat_id: int | None):
    name = name.strip()
    row = conn.execute(
        "SELECT id FROM ingredient WHERE name = ?", (name,)
    ).fetchone()
    if row:
        if cat_id is not None:
            conn.execute(
                "UPDATE ingredient SET category_id = COALESCE(category_id, ?), "
                "default_unit = COALESCE(default_unit, ?) WHERE id = ?",
                (cat_id, unit, row["id"]),
            )
        return row["id"]
    cur = conn.execute(
        "INSERT INTO ingredient (name, category_id, default_unit, shop) "
        "VALUES (?, ?, ?, NULL)",
        (name, cat_id, unit),
    )
    return cur.lastrowid


def _norm_item(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"\s+", " ", s)
    # drop quantity/pack chatter
    for junk in (
        "original",
        "geschnitten",
        "scheiben",
        "gemahlen",
        "glatt",
        "griffig",
        "billig",
        "groß",
        "grosse",
        "je ",
    ):
        s = s.replace(junk, " ")
    return re.sub(r"[^a-zäöüß0-9 ]", " ", s).strip()


# shopping_done labels → ingredient.name
SHOP_ALIASES = {
    "passata": "Tomaten Passiert",
    "tomatenpulpe": "Tomaten Passiert",
    "polpa": "Tomaten Gewürfelt",
    "tomatenfruchtfleisch": "Tomaten Gewürfelt",
    "sonnenblumenöl": "Öl",
    "balsamico": "Essig",
    "feinkristallzucker": "Zucker",
    "mehl glatt": "Mehl",
    "mehl griffig": "Mehl",
    "cremohonig": "Cremehonig",
    "creme-honig": "Cremehonig",
    "cini-minis": "CiniMinis",
    "cini minis": "CiniMinis",
    "erdbeermarmelade": "Marmelade",
    "marillenmarmelade": "Marmelade",
    "preiselbeermarmelade": "Marmelade",
    "gemüsebrühe": "Gemüsesuppe",
    "äpfel": "Jausenäpfel",
    "salat eisberg": "Salat",
    "erdäpfel zum grillen": "Kartoffeln",
    "erdäpfel speckig": "Kartoffeln",
    "putenbrust für geschnetzeltes": "Putenbrust",
    "putenbrust für schnitzel": "Putenbrust",
    "faschiertes für fleischlaibchen": "Faschiertes",
    "haschee": "Wursthaschee",
    "gouda geschnitten": "Gouda",
    "extrawurst geschnitten": "Extrawurst",
    "semmeln": "Semmerl",
    "muskatnuss": "Muskat-Nuss",
    "knoblauch": "Knoblauchgranulat",
    "senf kremser": "Senf",
    "currypuilver": "Currypulver",
    "currypulver": "Currypulver",
    "nektarinen/pfirsiche": "Nektarinen",
    "weintrauben blau u weiß": "Weintrauben",
    "käse gerieben": "Käse Gerieben",
}


def _match_ingredient_name(item: str, ingredient_names: list[str]) -> str | None:
    raw = item.strip()
    if not raw:
        return None
    # exact
    for n in ingredient_names:
        if n.lower() == raw.lower():
            return n
    norm = _norm_item(raw)
    for alias, target in SHOP_ALIASES.items():
        if alias in norm or norm in alias:
            if target in ingredient_names:
                return target
    # substring: ingredient name inside shopping item
    best = None
    best_len = 0
    for n in ingredient_names:
        nn = _norm_item(n)
        if len(nn) < 3:
            continue
        if nn in norm or norm in nn:
            if len(nn) > best_len:
                best, best_len = n, len(nn)
    return best


def _apply_shops_from_extras(conn) -> None:
    """Copy preferred shop from shopping_extra onto matching ingredients."""
    names = [r["name"] for r in conn.execute("SELECT name FROM ingredient")]
    matched = 0
    for row in conn.execute(
        """
        SELECT item, store FROM shopping_extra
        WHERE store IS NOT NULL AND TRIM(store) != ''
        """
    ):
        target = _match_ingredient_name(row["item"], names)
        if not target:
            continue
        conn.execute(
            """
            UPDATE ingredient
            SET shop = COALESCE(shop, ?)
            WHERE name = ?
            """,
            (row["store"].strip(), target),
        )
        matched += 1
    # sensible defaults for recipe ingredients still missing a shop
    defaults = {
        "Metro": [
            "Milch",
            "Eier",
            "Butter",
            "Joghurt",
            "Schlagobers",
            "Sauerrahm",
            "Haloumi",
            "Parmesan",
            "Käse Gerieben",
            "Spaghetti",
            "Fleckerl",
            "Spätzle",
            "Suppennudeln",
            "Kartoffelpürree",
            "Hascheeknödel",
            "Speckknödel",
            "Grammelknödel",
            "Spinatknödel",
            "Marillenknödel",
            "Erdbeerknödel",
            "Nougatknödel",
            "Gemüselaibchen",
            "Mischgemüse",
            "Erbsen",
            "Gulaschsaft",
            "Vegane Schnitzel",
            "Käsegriller",
            "Frankfurter",
            "Bratwürstel",
            "Käsekrainer",
            "Bernerwürstel",
            "Sellerie",
            "Schnittlauch",
            "Maizena",
            "Choco-Chips",
            "Tafelessig",
            "Pfeffer",
            "Salz",
            "Olivenöl",
            "Kräuter",
            "Zimt",
            "Majoran",
            "Oregano",
            "Thymian",
            "Basilikum",
            "Dill",
            "Tomatenmark",
            "Tomaten Passiert",
            "Tomaten Gewürfelt",
            "Essiggurkerl",
            "Kraut",
            "Paprika",
            "Mais",
            "Zucchini",
            "Zwiebel",
            "Karotten",
            "Kartoffeln",
            "Salat",
            "Salatgurken",
            "Champignons",
            "Bananen",
            "Weintrauben",
            "Nektarinen",
            "Jausenäpfel",
            "Putenbrust",
            "Faschiertes",
            "Wursthaschee",
            "Extrawurst",
            "Gouda",
            "Brot",
            "Mehl",
            "Zucker",
            "Staubzucker",
            "Kakao",
            "Kaffee",
            "Nutella",
            "Marmelade",
            "Backerbsen",
            "Reis",
            "Trockengerm",
            "Öl",
            "Ketchup",
            "Senf",
            "Currypulver",
            "Gemüsesuppe",
        ],
        "Billa": ["CiniMinis", "Weckerl"],
        "Spende Honeder": ["Knödelbrot", "Semmelbrösel"],
        "Bäcker vor Ort": ["Semmerl"],
    }
    for shop, items in defaults.items():
        for name in items:
            conn.execute(
                "UPDATE ingredient SET shop = ? WHERE name = ? AND shop IS NULL",
                (shop, name),
            )
    missing = conn.execute(
        "SELECT COUNT(*) FROM ingredient WHERE shop IS NULL"
    ).fetchone()[0]
    print(f"Shops matched from shopping_done (~{matched} hits); "
          f"{missing} ingredients still without shop")


def _apply_veggie_fixes(conn) -> None:
    """Mark veggie-only alternatives and set per-vegetarian amounts."""
    for recipe, ingredient, per_veg, total, note in VEGGIE_LINE_FIXES:
        cur = conn.execute(
            """
            UPDATE recipe_line
            SET audience = 'veggie',
                amount_per_person = ?,
                recipe_amount = ?,
                yield_portions = 2,
                notes = CASE
                  WHEN notes IS NULL OR notes = '' THEN ?
                  ELSE notes || ' | ' || ?
                END
            WHERE id IN (
              SELECT rl.id
              FROM recipe_line rl
              JOIN recipe r ON r.id = rl.recipe_id
              JOIN ingredient i ON i.id = rl.ingredient_id
              WHERE r.name = ? AND i.name = ?
            )
            """,
            (per_veg, total, note, note, recipe, ingredient),
        )
        if cur.rowcount == 0:
            print(f"WARNING: veggie fix not applied: {recipe} / {ingredient}")


def _apply_batch_overrides(conn) -> None:
    for recipe, ingredient, amount, note in BATCH_OVERRIDES:
        cur = conn.execute(
            """
            UPDATE recipe_line
            SET recipe_amount = ?,
                amount_per_person = ? / yield_portions,
                notes = CASE
                  WHEN notes IS NULL OR notes = '' THEN ?
                  WHEN notes LIKE '%' || ? || '%' THEN notes
                  ELSE notes || ' | ' || ?
                END
            WHERE id IN (
              SELECT rl.id FROM recipe_line rl
              JOIN recipe r ON r.id = rl.recipe_id
              JOIN ingredient i ON i.id = rl.ingredient_id
              WHERE r.name = ? AND i.name = ?
            )
            """,
            (amount, amount, note, note, note, recipe, ingredient),
        )
        if cur.rowcount == 0:
            print(f"WARNING: batch override not applied: {recipe} / {ingredient}")


def _apply_zero_fills(conn) -> None:
    """Replace placeholder zero lines with sensible batch amounts."""
    for recipe, ingredient, unit, amount, note in ZERO_FILLS:
        cur = conn.execute(
            """
            UPDATE recipe_line
            SET unit = ?,
                recipe_amount = ?,
                amount_per_person = ? / yield_portions,
                notes = CASE
                  WHEN notes IS NULL OR notes = '' THEN ?
                  WHEN notes LIKE '%' || ? || '%' THEN notes
                  ELSE notes || ' | ' || ?
                END
            WHERE id IN (
              SELECT rl.id FROM recipe_line rl
              JOIN recipe r ON r.id = rl.recipe_id
              JOIN ingredient i ON i.id = rl.ingredient_id
              WHERE r.name = ? AND i.name = ?
            )
            """,
            (unit, amount, amount, note, note, note, recipe, ingredient),
        )
        if cur.rowcount == 0:
            print(f"WARNING: zero-fill not applied: {recipe} / {ingredient}")
        else:
            conn.execute(
                """
                UPDATE ingredient
                SET default_unit = COALESCE(default_unit, ?)
                WHERE name = ?
                """,
                (unit, ingredient),
            )


def _apply_unit_rewrites(conn) -> None:
    """Clarify ambiguous pkg amounts (tomatoes, kraut, haloumi, …)."""
    for recipe, ingredient, unit, batch, yield_p, audience, note in UNIT_REWRITES:
        per = batch / yield_p
        cur = conn.execute(
            """
            UPDATE recipe_line
            SET unit = ?,
                recipe_amount = ?,
                yield_portions = ?,
                amount_per_person = ?,
                audience = ?,
                notes = CASE
                  WHEN notes IS NULL OR notes = '' THEN ?
                  WHEN notes LIKE '%' || ? || '%' THEN notes
                  ELSE notes || ' | ' || ?
                END
            WHERE id IN (
              SELECT rl.id FROM recipe_line rl
              JOIN recipe r ON r.id = rl.recipe_id
              JOIN ingredient i ON i.id = rl.ingredient_id
              WHERE r.name = ? AND i.name = ?
            )
            """,
            (unit, batch, yield_p, per, audience, note, note, note, recipe, ingredient),
        )
        if cur.rowcount == 0:
            print(f"WARNING: unit rewrite not applied: {recipe} / {ingredient}")
        else:
            conn.execute(
                "UPDATE ingredient SET default_unit = ? WHERE name = ?",
                (unit, ingredient),
            )


def _meal_slot_id(conn, camp_id: int, day: str, meal: str) -> int | None:
    row = conn.execute(
        """
        SELECT id FROM meal_slot
        WHERE camp_id = ? AND day_name = ? AND meal = ?
        """,
        (camp_id, day, meal),
    ).fetchone()
    return int(row["id"]) if row else None


def _set_meal_recipes(conn, slot_id: int, recipes: list[str]) -> None:
    conn.execute("DELETE FROM meal_recipe WHERE meal_slot_id = ?", (slot_id,))
    for i, name in enumerate(recipes):
        row = conn.execute("SELECT id FROM recipe WHERE name = ?", (name,)).fetchone()
        if not row:
            print(f"WARNING: recipe missing for menu swap: {name}")
            continue
        conn.execute(
            """
            INSERT INTO meal_recipe (meal_slot_id, recipe_id, sort_order)
            VALUES (?, ?, ?)
            """,
            (slot_id, row["id"], i),
        )


def _apply_evening_menu_swaps(conn, camp_id: int) -> None:
    """2026: Knödel Mon→Thu, Palatschinken Tue→Mon (no Nudelsuppe), Fleckerlspeise Thu→Tue."""
    swaps = [
        (
            "Montag",
            "Abend",
            ["Palatschinken"],
            1,
            "palatschinken, knacker + smores",
            "mehl (susi)",
        ),
        (
            "Dienstag",
            "Abend",
            ["Fleckerlspeise", "Gurkensalat", "Karottensalat"],
            1,
            "haschee + schinken",
            "nudeln (susi)",
        ),
        (
            "Donnerstag",
            "Abend",
            ["Knödel"],
            0,
            None,
            "semmelknödel (susi bringt)",
        ),
    ]
    for day, meal, recipes, veggie, notes, gluten in swaps:
        slot_id = _meal_slot_id(conn, camp_id, day, meal)
        if slot_id is None:
            print(f"WARNING: meal slot missing for swap: {day} {meal}")
            continue
        conn.execute(
            """
            UPDATE meal_slot
            SET veggie_option = ?, notes = ?, gluten_notes = ?
            WHERE id = ?
            """,
            (veggie, notes, gluten, slot_id),
        )
        _set_meal_recipes(conn, slot_id, recipes)


def _apply_menu_overrides(conn, camp_id: int) -> None:
    """Drop Zucchinicremesuppe — note 'keine suppe 2025'."""
    slot = conn.execute(
        """
        SELECT id FROM meal_slot
        WHERE camp_id = ? AND day_name = 'Sonntag' AND meal = 'Abend'
        """,
        (camp_id,),
    ).fetchone()
    if not slot:
        return
    conn.execute(
        """
        DELETE FROM meal_recipe
        WHERE meal_slot_id = ?
          AND recipe_id = (SELECT id FROM recipe WHERE name = 'Zucchinicremesuppe')
        """,
        (slot["id"],),
    )
    conn.execute(
        """
        UPDATE meal_slot
        SET notes = CASE
          WHEN notes IS NULL OR notes = '' THEN 'Zucchinicremesuppe gestrichen (keine suppe 2025)'
          WHEN notes LIKE '%keine suppe 2025%' THEN notes
          ELSE notes || ' | Zucchinicremesuppe gestrichen (keine suppe 2025)'
        END
        WHERE id = ?
        """,
        (slot["id"],),
    )


def import_xlsx(xlsx_path: Path, db_path: Path | None = None) -> Path:
    wb = load_workbook(xlsx_path, data_only=False)
    conn = connect(db_path)
    # fresh schema
    for table in (
        "shopping_extra",
        "meal_recipe",
        "meal_slot",
        "recipe_line",
        "recipe",
        "ingredient",
        "category",
        "attendance_group",
        "camp",
    ):
        conn.execute(f"DROP TABLE IF EXISTS {table}")
    init_db(conn)

    cur = conn.execute(
        "INSERT INTO camp (name, year, start_date, notes) VALUES (?, ?, ?, ?)",
        (
            "Jungscharlager",
            2026,
            "2026-08-30",
            "Imported from küchenplan_26.xlsx; Kaiserschmarrn → Palatschinken",
        ),
    )
    camp_id = cur.lastrowid

    # --- categories from categoryMap headers ---
    cm = wb["categoryMap"]
    headers = []
    for c in range(1, cm.max_column + 1):
        h = cm.cell(1, c).value
        if c == 1:
            h = "Non Food"
        headers.append(h)
    cat_ids: dict[str, int] = {}
    for i, name in enumerate(headers):
        if not name:
            continue
        cur = conn.execute(
            "INSERT INTO category (name, sort_order) VALUES (?, ?)", (name, i)
        )
        cat_ids[name] = cur.lastrowid

    # ingredient → category
    name_to_cat: dict[str, int] = {}
    for c, cat in enumerate(headers, 1):
        if not cat or cat not in cat_ids:
            continue
        for r in range(2, cm.max_row + 1):
            v = cm.cell(r, c).value
            if v:
                name_to_cat[str(v).strip()] = cat_ids[cat]

    # Marmelade not in map — put with other jams
    if "Marmelade" not in name_to_cat and "Frühstück, Backen" in cat_ids:
        name_to_cat["Marmelade"] = cat_ids["Frühstück, Backen"]

    # --- attendance (weekPlan J2:L3 factors from J4:L4 formulas) ---
    wp = wb["weekPlan"]
    # leiter
    attendance = [
        ("leiter", "normal", _num(wp["J2"].value) or 0.0, 1.2),
        ("leiter", "strong_eater", _num(wp["K2"].value) or 0.0, 1.5),
        ("leiter", "veggie", _num(wp["L2"].value) or 0.0, 1.2),
        ("kinder", "normal", _num(wp["J3"].value) or 0.0, 0.8),
        ("kinder", "strong_eater", _num(wp["K3"].value) or 0.0, 1.2),
        ("kinder", "veggie", _num(wp["L3"].value) or 0.0, 0.8),
    ]
    for label, etype, count, factor in attendance:
        conn.execute(
            """
            INSERT INTO attendance_group (camp_id, label, eater_type, count, factor)
            VALUES (?, ?, ?, ?, ?)
            """,
            (camp_id, label, etype, count, factor),
        )

    # --- recipes ---
    rs = wb["recipes"]
    recipe_ids: dict[str, int] = {}
    for r in range(2, rs.max_row + 1):
        rname = rs.cell(r, 1).value
        if not rname:
            continue
        rname = str(rname).strip()
        if rname == "Kaiserschmarrn":
            continue  # replaced below

        ing = rs.cell(r, 2).value
        if not ing:
            continue
        ing = str(ing).strip()
        unit = str(rs.cell(r, 3).value or "").strip() or "stk"
        amount = _num(rs.cell(r, 4).value) or 0.0
        portions = _num(rs.cell(r, 5).value) or 40.0
        per = _num(rs.cell(r, 6).value)
        raw_i = rs.cell(r, 9).value
        per_final = _num(_cached_dummy(raw_i) if isinstance(raw_i, str) else raw_i)
        if per is None:
            per = amount / portions if portions else 0.0
        if per_final is None:
            per_final = per
        zus = rs.cell(r, 7).value
        is_side = 1 if zus in (1, "1", True) else 0
        notes = rs.cell(r, 8).value
        adj = rs.cell(r, 10).value
        note_bits = [x for x in (notes, adj) if x]
        note = " | ".join(str(x) for x in note_bits) or None

        if rname not in recipe_ids:
            cur = conn.execute(
                "INSERT INTO recipe (name, notes, source) VALUES (?, ?, ?)",
                (rname, None, "küchenplan_26.xlsx"),
            )
            recipe_ids[rname] = cur.lastrowid

        cat_id = name_to_cat.get(ing)
        iid = _get_or_create_ingredient(conn, ing, unit, cat_id)
        conn.execute(
            """
            INSERT INTO recipe_line (
              recipe_id, ingredient_id, unit, recipe_amount, yield_portions,
              amount_per_person, is_side, audience, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'all', ?)
            """,
            (
                recipe_ids[rname],
                iid,
                unit,
                amount,
                portions,
                float(per_final),
                is_side,
                note,
            ),
        )

    # --- Palatschinken ---
    cur = conn.execute(
        "INSERT INTO recipe (name, notes, source) VALUES (?, ?, ?)",
        (
            "Palatschinken",
            "Scaled from 5 → 40 portions (×8)",
            PALATSCHINKEN_SOURCE,
        ),
    )
    pala_id = cur.lastrowid
    recipe_ids["Palatschinken"] = pala_id
    for ing, unit, amount40, note in PALATSCHINKEN_LINES:
        cat_id = name_to_cat.get(ing)
        iid = _get_or_create_ingredient(conn, ing, unit, cat_id)
        per = amount40 / 40.0
        conn.execute(
            """
            INSERT INTO recipe_line (
              recipe_id, ingredient_id, unit, recipe_amount, yield_portions,
              amount_per_person, is_side, audience, notes
            ) VALUES (?, ?, ?, ?, 40, ?, 0, 'all', ?)
            """,
            (pala_id, iid, unit, amount40, per, note),
        )

    _apply_veggie_fixes(conn)
    _apply_batch_overrides(conn)
    _apply_zero_fills(conn)
    _apply_unit_rewrites(conn)

    # --- meal plan ---
    for r in range(2, wp.max_row + 1):
        day = wp.cell(r, 1).value
        meal = wp.cell(r, 2).value
        recipes_cell = wp.cell(r, 3).value
        if not day or not meal or not recipes_cell:
            continue
        day = str(day).strip()
        meal = str(meal).strip()
        names = _split_recipes(recipes_cell)
        names = [
            "Palatschinken" if n == "Kaiserschmarrn" else n for n in names
        ]
        cur = conn.execute(
            """
            INSERT INTO meal_slot (
              camp_id, day_name, day_index, meal, headcount_note,
              veggie_option, notes, gluten_notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                camp_id,
                day,
                DAY_INDEX.get(day, 99),
                meal,
                wp.cell(r, 4).value,
                1 if wp.cell(r, 5).value else 0,
                wp.cell(r, 6).value,
                wp.cell(r, 7).value,
            ),
        )
        slot_id = cur.lastrowid
        for i, n in enumerate(names):
            rid = recipe_ids.get(n)
            if rid is None:
                # create empty recipe stub so menu stays intact
                cur = conn.execute(
                    "INSERT INTO recipe (name, notes, source) VALUES (?, ?, ?)",
                    (n, "Referenced on menu but missing from recipes sheet", None),
                )
                rid = cur.lastrowid
                recipe_ids[n] = rid
            conn.execute(
                """
                INSERT INTO meal_recipe (meal_slot_id, recipe_id, sort_order)
                VALUES (?, ?, ?)
                """,
                (slot_id, rid, i),
            )

    _apply_menu_overrides(conn, camp_id)
    _apply_evening_menu_swaps(conn, camp_id)

    # --- shopping_done extras ---
    sd = wb["shopping_done"]
    current_cat = None
    for r in range(1, sd.max_row + 1):
        a, b, c, d = (sd.cell(r, col).value for col in range(1, 5))
        if a and not b and not c:
            current_cat = str(a).strip()
            continue
        if not b:
            continue
        item = str(b).strip()
        if item.lower().startswith("shopping"):
            continue
        conn.execute(
            """
            INSERT INTO shopping_extra (camp_id, category, quantity_text, item, note, store)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                camp_id,
                current_cat,
                str(a) if a is not None else None,
                item,
                str(c) if c is not None else None,
                str(d) if d is not None else None,
            ),
        )

    _apply_shops_from_extras(conn)

    conn.commit()
    path = Path(conn.execute("PRAGMA database_list").fetchone()[2])
    portions = portion_equivalents(conn, camp_id)
    n_recipes = conn.execute("SELECT COUNT(*) FROM recipe").fetchone()[0]
    n_lines = conn.execute("SELECT COUNT(*) FROM recipe_line").fetchone()[0]
    n_meals = conn.execute("SELECT COUNT(*) FROM meal_slot").fetchone()[0]
    conn.close()
    print(f"DB: {path}")
    print(f"Recipes: {n_recipes}, lines: {n_lines}, meal slots: {n_meals}")
    print(f"Portion equivalents: {portions}")
    return path


def main(argv: list[str] | None = None) -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "xlsx",
        nargs="?",
        default=str(Path.home() / "Downloads" / "küchenplan_26.xlsx"),
    )
    p.add_argument("--db", default=None)
    args = p.parse_args(argv)
    from . import DEFAULT_DB

    import_xlsx(Path(args.xlsx), Path(args.db) if args.db else DEFAULT_DB)


if __name__ == "__main__":
    main()
